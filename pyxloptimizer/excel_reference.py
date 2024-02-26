import ast
import re
from enum import StrEnum
from collections import namedtuple

import openpyxl
import numpy

from .logger import logger
from .shape import Shape
from .exceptions import InvalidExcelReferenceException

ROW_FORMAT_EXPRESSION = r"^\d+:\d+$"
COLUMN_FORMAT_EXPRESSION = r"^[A-Z]{1,3}:[A-Z]{1,3}$"
RANGE_FORMAT_EXPRESSION = r"^[A-Z]{0,3}\d+:[A-Z]{1,3}\d+$"
ARRAY_REF_EXPRESSION = r"^[A-Z]{0,3}\d+#$"
CELL_FORMAT_EXPRESSION = r"^[A-Z]{0,3}\d+$"

# Inspired by Pycel precedents of operators.
# https://support.microsoft.com/en-us/office/
#       calculation-operators-and-precedence-in-excel-48be406d-4975-4d31-b2b8-7af9e0e2878a
BinOperatorDetails = namedtuple('BinOperatorDetails', ['precedence', 'astOperator', 'comparison'])
BIN_OP_MAP = {
    '^': BinOperatorDetails(5, ast.Pow, False),
    '*': BinOperatorDetails(4, ast.Mult, False),
    '/': BinOperatorDetails(4, ast.Div, False),
    '+': BinOperatorDetails(3, ast.Add, False),
    '-': BinOperatorDetails(3, ast.Sub, False),
    '&': BinOperatorDetails(2, ast.Add, False),  # need to special case this to ensure we switch everything to strings
    '=': BinOperatorDetails(1, ast.Eq, True),
    '<': BinOperatorDetails(1, ast.Lt, True),
    '>': BinOperatorDetails(1, ast.Gt, True),
    '<=': BinOperatorDetails(1, ast.LtE, True),
    '>=': BinOperatorDetails(1, ast.GtE, True),
    '<>': BinOperatorDetails(1, ast.NotEq, True),
}

RangeBoundaries = namedtuple("RangeBoundaries", ["min_col", "min_row", "max_col", "max_row"])


class SheetDetails:
    def __init__(self, sheet):
        self._arrays = {}
        for src_ref, array_ref in sheet.array_formulae.items():
            a_range = get_cell_range(array_ref)
            shape = Shape(a_range.max_row - a_range.min_row + 1, a_range.max_col - a_range.min_col + 1)

            if shape.is_scalar:
                # single cell array references are a bit of an oddity and not useful for our purposes.
                continue
            else:
                # todo there is a merge request in openpyxl to support the metadata on array functions
                # https://foss.heptapod.net/openpyxl/openpyxl/-/merge_requests/439
                # eventually should integrate here to have proper behvaiour for rand, row functions
                # when used in array context.
                self._arrays[a_range] = src_ref

    def get_array_src(self, cell_range):
        """
        Determien if this cell range is a portion of an array;
        Todo - this in O(n) so might be a problem if a sheet has a lot of array formulas
        Todo - if the cell refernece _partially_ overlaps with arrays it is more complex; and not currently supported.
        but that seems unlikely.
        """
        result = None
        for array_range, array_src in self._arrays.items():
            # Todo figure out non-overlapping region.
            if cell_range.min_col >= array_range.min_col and cell_range.max_col <= array_range.max_col and \
                    cell_range.min_row >= array_range.min_row and cell_range.max_row <= array_range.max_row:
                result = array_src
        return result


class ExcelReference:
    class ReferenceType(StrEnum):
        CELL = 'cell'
        RANGE = 'range'
        ARRAY = 'array_entry'
        ARRAY_REF = 'array_reference'

    def __init__(self, wb: openpyxl.Workbook, address: str, active_sheet: str = None):
        self._wb = wb
        if active_sheet is None:
            active_sheet = wb.active.title

        if not hasattr(self._wb[active_sheet], '__xlnumba'):
            wb_details = SheetDetails(self._wb[active_sheet])
            setattr(self._wb[active_sheet], '__xlnumba', wb_details)
        else:
            wb_details = getattr(self._wb[active_sheet], '__xlnumba')

        self._sheet, self._cell_ref = _split_cell_address(address, active_sheet)

        if re.match(COLUMN_FORMAT_EXPRESSION, self._cell_ref):
            col = self._cell_ref.split(":")[0]
            self._cell_ref = f"{col}1:{col}{wb[self._sheet].max_row}"
        elif re.match(ROW_FORMAT_EXPRESSION, self._cell_ref):
            row = self._cell_ref.split(":")[1]
            max_col = openpyxl.utils.cell.get_column_letter(wb[self._sheet].max_column)
            self._cell_ref = f"A{row}:{max_col}{row}"

        self.range_name = None
        if self._cell_ref in self._wb.defined_names:
            # If this is a defined range we demask it for compilation purposes we don't need to know that
            # however, for output and other purposes useful to keep track of.
            self._range_name = self._cell_ref
            array_range = wb.defined_names[self._cell_ref]
            self._sheet, self._cell_ref = _split_cell_address(array_range.value, active_sheet)

        self._compute_ref_type()

        # confirm sheet actually exists in workbook
        if self._sheet not in self._wb.sheetnames:
            raise InvalidExcelReferenceException(f"Sheet {self._sheet} does not exist in workbook")

        self._array_src = None
        if self._ref_type != self.ReferenceType.ARRAY_REF:
            cell_range = get_cell_range(self._cell_ref)
            src_ref = wb_details.get_array_src(cell_range)
            if src_ref is not None:
                array_range = self._wb[self._sheet].array_formulae.get(src_ref)
                self._array_src = src_ref + '#'

                if cell_range == array_range:
                    logger.debug(f"Found range array reference {src_ref}# for range {self._cell_ref}")
                    assert self._ref_type != self.ReferenceType.CELL
                    self._ref_type = self.ReferenceType.ARRAY_REF
                else:
                    self._ref_type = self.ReferenceType.ARRAY

    def create_relative(self, coord):
        """ Create an Excel reference on the same sheet as current reference (if sheet not specified)"""
        return ExcelReference(self._wb, coord, self._sheet)

    @property
    def array_src(self) -> 'ExcelReference':
        assert self.is_array
        return ExcelReference(self._wb, self._array_src, self._sheet)

    @property
    def offsets(self):
        assert self.is_array
        array_range = self._wb[self._sheet].array_formulae[self._array_src[:-1]]
        offset = compute_contained_offsets(self._cell_ref, array_range)
        return offset

    @property
    def is_range(self) -> bool:
        return self._ref_type == self.ReferenceType.RANGE

    @property
    def is_array(self) -> bool:
        return self._ref_type == self.ReferenceType.ARRAY

    @property
    def row(self) -> int | numpy.ndarray:
        """
        Row computes the row value of a reference; if this is scalar reference is easy
        """
        cell_ref = self._cell_ref
        if self._ref_type == self.ReferenceType.ARRAY_REF:
            assert self._cell_ref.endswith("#")
            cell_ref = self._cell_ref[:-1]
        return self._wb[self._sheet][cell_ref].row

    @property
    def column(self) -> int:
        return self._wb[self._sheet][self._cell_ref].col_idx

    @property
    def sheet(self) -> str:
        return self._sheet

    @property
    def sheet_count(self) -> int:
        return len(self._wb.sheetnames)

    @property
    def shape(self) -> Shape:
        if self._ref_type == self.ReferenceType.RANGE or self._ref_type == self.ReferenceType.ARRAY:
            min_col, min_row, max_col, max_row = openpyxl.utils.cell.range_boundaries(self._cell_ref)
            return Shape(max_row - min_row + 1, max_col - min_col + 1)
        elif self._ref_type == self.ReferenceType.ARRAY_REF:
            assert self._cell_ref[-1] == "#"
            tmp_ref = self._wb[self._sheet].array_formulae[self._cell_ref[:-1]]
            min_col, min_row, max_col, max_row = openpyxl.utils.cell.range_boundaries(tmp_ref)
            return Shape(max_row - min_row + 1, max_col - min_col + 1)
        else:
            return Shape(1, 1)

    def get_cells(self):
        """
        For range and aray type objects return the cells that underly this object.
        """
        if self._ref_type == self.ReferenceType.RANGE:
            # todo fill this in!
            min_col, min_row, max_col, max_row = openpyxl.utils.cell.range_boundaries(self._cell_ref)

            vals = []
            for i in range(min_row, max_row + 1):
                for j in range(min_col, max_col + 1):
                    vals.append(
                        ExcelReference(self._wb,
                                       f"{self._sheet}!{openpyxl.utils.cell.get_column_letter(j)}{i}",
                                       self._sheet))
            return vals
        else:
            raise NotImplementedError()

    def encode_name(self):
        """ Craete a string usable as a variable name from this Excel reference. """
        return (self._sheet + "_" + self._cell_ref.replace(":", "x")).replace("#", "_xx").lower()

    def _compute_ref_type(self):
        """ Part of init, this determines the type of reference we have based on format """
        if re.match(CELL_FORMAT_EXPRESSION, self._cell_ref):
            self._ref_type = self.ReferenceType.CELL
        elif re.match(RANGE_FORMAT_EXPRESSION, self._cell_ref):
            self._ref_type = self.ReferenceType.RANGE
        elif re.match(ARRAY_REF_EXPRESSION, self._cell_ref):
            self._ref_type = self.ReferenceType.ARRAY_REF
        else:
            raise InvalidExcelReferenceException(f"Invalid cell reference {self._cell_ref}")

    @property
    def value(self):
        if self._ref_type == ExcelReference.ReferenceType.ARRAY_REF:
            assert self._cell_ref[-1] == '#'
            cell = self._wb[self._sheet][self._cell_ref[:-1]]
            assert cell.value.t == 'array'
            return cell.value.text
        else:
            assert self._ref_type == ExcelReference.ReferenceType.CELL
            cell = self._wb[self._sheet][self._cell_ref]
            cell_value = cell.value
            data_type = cell.data_type
            if cell_value is None:
                return None
            elif isinstance(cell_value, openpyxl.worksheet.formula.ArrayFormula):
                assert cell_value.t == 'array'
                return cell_value.text
            elif data_type == 'n':
                return cell_value
            elif data_type == 'b':
                return cell.value
            else:
                return str(cell.value)

    @property
    def data_type(self):
        if self._ref_type == self.ReferenceType.ARRAY_REF:
            return 'f'  # always a formula
        else:
            assert self._ref_type == self.ReferenceType.CELL
            if self._wb[self._sheet][self._cell_ref].value is None:
                return 'X'
            else:
                return self._wb[self._sheet][self._cell_ref].data_type

    def __eq__(self, other):
        assert isinstance(other, ExcelReference)
        return self._cell_ref == other._cell_ref and self._sheet == other._sheet

    def __repr__(self):
        return f"{self._sheet}!{self._cell_ref}"

    def __hash__(self):
        return hash(str(self))


def _split_cell_address(address: str, default_sheet=None):
    """
    Given a standard address in Excel (Sheet1!A1) split this into a sheet and address section,
    handling that sheets are defaulted if not specified and quotation if needed.
    :param address:  Address to split
    :param default_sheet: Sheet reference came from to determine this location.
    :return: both the sheet and cell reference for this address.
    """
    if "!" in address:
        # split by the last !, as earlier ones can be quoted.
        sheet, address = address.rsplit('!', maxsplit=1)

        if sheet.startswith("'"):
            assert sheet.endswith("'")  # ensure properly quoted or something happened we don't expect
            sheet = sheet[1:-1].replace("''", "'")
    else:
        # if the sheet is not specified (relative reference) then the default sheet must be passed in by calling
        # function
        assert default_sheet
        sheet = default_sheet

    # $ in cell ref are used to keep references constant when broadcast in Excel.  Since we are not modifying
    # the sheet we can safely remove them and simplify processing globally.
    address = address.replace("$", "")

    return sheet, address


def compute_contained_offsets(cell, array):
    cell_range = get_cell_range(cell)
    array_range = get_cell_range(array)

    return (
        (cell_range.min_row - array_range.min_row, cell_range.max_row - array_range.min_row + 1),
        (cell_range.min_col - array_range.min_col, cell_range.max_col - array_range.min_col + 1)
    )


def get_cell_range(excel_ref_str):
    # Let's figure out how
    tpls = openpyxl.utils.cell.range_boundaries(excel_ref_str)
    return RangeBoundaries(*tpls)


class DataType(StrEnum):
    """
    Enum for the return type of function.  This is used to determine how the function should be called.

    The codes match as much as possible with Excel to make it easier to remember, but should not
    map between two directly.
    """
    Number = 'n'
    String = 's'
    Boolean = 'b'
    Date = 'd'
    Blank = 'X'  # special case should only be the case for input data, not geenrated later.
