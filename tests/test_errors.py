import openpyxl
import pytest

from pyxloptimizer import Compiler, InvalidExcelReferenceException, UnsupportedException
from pyxloptimizer.excel_functions import find_function_details


@pytest.fixture(scope='module')
def xls_sheet(request):
    return openpyxl.load_workbook('tests/fixtures/basic.xlsx')


def test_read_only_sheet():
    with pytest.raises(AttributeError):
        wb = openpyxl.load_workbook('tests/fixtures/basic.xlsx', read_only=True)
        Compiler(wb)


def test_undefined_reference(xls_sheet):
    """
    Most basic of test cases - ensure that a simple formula can be processed, and produce a result
    given an input.
    """
    with pytest.raises(InvalidExcelReferenceException):
        ctx = Compiler(xls_sheet)
        ctx.add_input("src", "InputT")


def test_no_output(xls_sheet):
    """
    Test what happens if we don't have an output cell.
    """
    with pytest.raises(AttributeError):
        ctx = Compiler(xls_sheet)
        ctx.add_input("src", "A1")
        ctx.compile()


def test_invalid_sheet(xls_sheet):
    """
    Most basic of test cases - ensure that a simple formula can be processed, and produce a result
    given an input.
    """
    with pytest.raises(InvalidExcelReferenceException):
        ctx = Compiler(xls_sheet)
        ctx.add_input("src", "SheetX!A1")


def test_readonly_wb():
    with pytest.raises(AttributeError):
        xls_sheet = openpyxl.load_workbook('tests/fixtures/basic.xlsx', read_only=True)
        Compiler(xls_sheet)


def test_blank_array(xls_sheet):
    """
    Most basic of test cases - ensure that a simple formula can be processed, and produce a result
    given an input.
    """
    with pytest.raises(UnsupportedException):
        ctx = Compiler(xls_sheet)
        ctx.add_input("src", "Sheet1!A1")
        ctx.add_output("dst", "Sheet1!F1")
        fn = ctx.compile(disable_numba=True)
        fn(src=2)


def test_mixed_array_type(xls_sheet):
    """
    Most basic of test cases - ensure that a simple formula can be processed, and produce a result
    given an input.
    """
    with pytest.raises(UnsupportedException):
        ctx = Compiler(xls_sheet)
        ctx.add_input("src", "Sheet1!A11")
        ctx.add_output("dst", "Sheet1!A15")
        fn = ctx.compile(disable_numba=True)
        fn(src=2)


def test_notimplmented_functions():
    with pytest.raises(NotImplementedError):
        find_function_details("XX")


def test_unsupported_functions():
    with pytest.raises(UnsupportedException):
        find_function_details("T")
